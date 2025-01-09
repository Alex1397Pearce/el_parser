import pandas as pd
from selenium import webdriver
from selenium.common.exceptions import TimeoutException, NoSuchElementException, StaleElementReferenceException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import requests
import random
import time
import openpyxl
import os

# Настройки для Selenium
chrome_options = Options()
# chrome_options.add_argument('--headless')

# Функция для получения кода состояния
def get_status_code(driver):
    return driver.execute_script("""
        var xhr = new XMLHttpRequest();
        xhr.open('GET', '{}', false);
        xhr.send();
        return String(xhr.status);
    """.format(driver.current_url))

# Функция для чтения столбца из Excel
def read_excel_column(file_path, column_name):
    df = pd.read_excel(file_path)
    return df[column_name].tolist()

# Функция для получения страницы
def get_page(driver, value, url_search):
    print("Процедура: get_page.")
    try:
        url = url_search + value
        driver.get(url)
        time.sleep(10)
        status_code = get_status_code(driver)
        print(status_code, type(status_code))
        if status_code == "200":
            print("Страница успешно загружена.")
            element = WebDriverWait(driver, 5).until(
                EC.visibility_of_element_located((By.XPATH, '//div[@class="media-left"]/a'))
            )
            src_value = element.get_attribute('href')
            time.sleep(random.randrange(2, 6))
            return src_value
        else:
            print("Произошла ошибка при загрузке страницы.")
            return "status_code"
    except (TimeoutException, NoSuchElementException, StaleElementReferenceException, WebDriverException) as e:
        print(f"Произошла ошибка: {e}")
        print(f'Артикул {value} не найден')
        return 1

# Функция для получения ссылки на изображение
def get_picture(driver, url):
    print("Процедура: get_picture.")
    try:
        driver.get(url)
        time.sleep(10)
        status_code = get_status_code(driver)
        if status_code == "200":
            print("Страница успешно загружена.")
            element = WebDriverWait(driver, 10).until(
                EC.presence_of_element_located((By.XPATH, '//div[@class="big_image"]/a'))
            )
            src_value = element.get_attribute('href')
            print(f'Значение атрибута src: {src_value}')
            time.sleep(random.randrange(2, 6))
            return src_value
        else:
            print("Произошла ошибка при загрузке страницы.")
            return "status_code"
    except (TimeoutException, NoSuchElementException, StaleElementReferenceException, WebDriverException) as e:
        print(f"Ошибка в процедуре get_picture(). src_value = {url}")
        print(e)
        return None

# Функция для очистки Excel файла
def clean_excel(file_path):
    try:
        os.remove(file_path)
    except FileNotFoundError:
        print(f"Файл {file_path} не найден.")

# Функция для записи в Excel файл
def list_to_excel(value, status, path=""):
    file_path = r"\\1cdbsrv\SystemFiles\pictures\result\Succes_ERA.xlsx" if status == "Success" else r"\\1cdbsrv\SystemFiles\pictures\result\Failed_ERA.xlsx"
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

    workbook = openpyxl.load_workbook(file_path)
    while len(workbook.sheetnames) > 1:
        workbook.remove(workbook[workbook.sheetnames[1]])

    if not workbook.sheetnames:
        workbook.create_sheet()

    sheet = workbook.active
    first_empty_row = 1
    while sheet.cell(row=first_empty_row, column=1).value is not None:
        first_empty_row += 1

    sheet.cell(row=first_empty_row, column=1, value=value)
    if status == "Success":
        sheet.cell(row=first_empty_row, column=2, value=path)

    workbook.save(file_path)

# Функция для загрузки файла
def download(url, name_file, comment):
    print("Процедура: download.")
    response = requests.get(url, verify=False)
    name_file = fr"\\1cdbsrv\SystemFiles\pictures\{name_file}"
    if response.status_code == 200:
        with open(name_file, 'wb') as file:
            file.write(response.content)
        print(f'Файл {name_file} успешно загружен. Ссылка на товар - {comment}')
        return name_file
    else:
        print(f'Ошибка при загрузке файла. Статус код: {response.status_code}')
        return False

def main():
    # Объявление глобальных переменных
    excel_file_path = 'era.xlsx'
    excel_column_name = 'Артикул'
    url_search = 'https://www.eraworld.ru/search?q='

    excel_values = read_excel_column(excel_file_path, excel_column_name)
    # excel_values = read_excel_column(excel_file_path, excel_column_name)
    count_not = 0
    count_yes = 0
    clean_excel(file_path=r"\\1cdbsrv\SystemFiles\pictures\result\Succes_ERA.xlsx")
    clean_excel(file_path=r"\\1cdbsrv\SystemFiles\pictures\result\Failed_ERA.xlsx")

    with webdriver.Chrome(options=chrome_options) as driver:
        for value in excel_values:
            condition = False
            while not condition:
                url_tovar = get_page(driver, value, url_search)
                if url_tovar != "status_code":
                    break
                else:
                    driver.refresh()
            if url_tovar == 1:
                count_not += 1
                list_to_excel(value, "Fail")
            else:
                while not condition:
                    url_pic = get_picture(driver, url_tovar)
                    if url_pic != "status_code":
                        break
                    else:
                        driver.refresh()
                name_file = value + ".png"
                path = download(url_pic, name_file, url_tovar)
                count_yes += 1
                list_to_excel(value, "Success", path)

    print(f"Count_not = {count_not}")
    print(f"Count_yes = {count_yes}")

if __name__ == "__main__":
    main()
