import pandas as pd
from selenium import webdriver
from selenium.common import TimeoutException, NoSuchElementException, StaleElementReferenceException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import requests
import random
import time
import openpyxl
import os

# Объявление глобальных переменных
excel_file_path = 'ekf.xlsx'
excel_column_name = 'Артикул'
url_search = 'https://ekfgroup.com/ru/search?q='

chrome_options = Options()
# chrome_options.add_argument('--headless')
driver = webdriver.Chrome(options=chrome_options)

def read_excel_column(file_path, column_name):
    if not file_path or not column_name:
        print("File path or column name is null.")
        return []

    # Чтение определенного столбца из Excel файла и возвращение его в виде списка
    try:
        df = pd.read_excel(file_path)
        column_values = df[column_name].tolist()
        return column_values
    except Exception as e:
        print(f"Error reading Excel file: {e}")
        return []

def is_url(input_string):
    if not input_string:
        return False

    try:
        result = requests.get(input_string)
        return result.status_code
    except ValueError:
        return False

def GetPage(value):
    if not value:
        print("Value is null.")
        return 1

    try:
        url = url_search + value
        driver.get(url)
        parent_element = driver.find_elements(By.CSS_SELECTOR, 'a.text-body.text-hover-primary.text-decoration-none')
        return parent_element
    except TimeoutException:
        print(f'Произошла ошибка ожидания. Артикул {value} не найден.')
        return 1
    except NoSuchElementException:
        print(f'Элемент не был найден. Артикул {value} не найден.')
        return 1
    except StaleElementReferenceException:
        print(f'Произошла ошибка с устаревшим элементом. Артикул {value} не найден.')
        return 1
    except WebDriverException as e:
        print(f'Произошла общая ошибка WebDriver. Артикул {value} не найден. Ошибка: {e}')
        return 1

def GetPicture(url):
    if not url:
        print("URL is null.")
        return None

    try:
        driver.get(url)
        element = WebDriverWait(driver, 5).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'gallery-slide-image'))
        )
        src_value = element.get_attribute('src')
        print(f'Значение атрибута src: {src_value}')
        wait = random.randrange(2, 6)
        time.sleep(wait)
        return src_value
    except TimeoutException:
        print(f"Ошибка в процедуре GetPicture(). URL: {url}")
        return None
    except Exception as e:
        print(f"Ошибка в процедуре GetPicture(): {e}")
        return None

def cleanExcel(file_path):
    if not file_path:
        print("File path is null.")
        return

    try:
        os.remove(file_path)
    except FileNotFoundError:
        print(f"Файл {file_path} не найден.")

def listToExcel(value, status):
    if not value or not status:
        print("Value or status is null.")
        return

    file_path = r"\\1cdbsrv\SystemFiles\pictures\result\Succes_EKF.xlsx" if status == "Success" else r"\\1cdbsrv\SystemFiles\pictures\result\Failed_EKF.xlsx"
    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

    workbook = openpyxl.load_workbook(file_path)
    while len(workbook.sheetnames) > 1:
        workbook.remove(workbook[workbook.sheetnames[1]])

    if not workbook.sheetnames:
        workbook.create_sheet()

    sheet = workbook.active
    first_empty_row = 1
    while sheet.cell(row=first_empty_row, column=1).value is not None:
        first_empty_row += 1

    sheet.cell(row=first_empty_row, column=1, value=value)
    workbook.save(file_path)

def download(url, name_file, comment):
    if not url or not name_file or not comment:
        print("URL, name_file or comment is null.")
        return False

    try:
        response = requests.get(url, verify=False)
        name_file = fr"\\1cdbsrv\SystemFiles\pictures\{name_file}"
        if response.status_code == 200:
            with open(name_file, 'wb') as file:
                file.write(response.content)
            print(f'Файл {name_file} успешно загружен. Ссылка на товар - {comment}')
            return True
        else:
            print(f'Ошибка при загрузке файла. Статус код: {response.status_code}')
            return False
    except Exception as e:
        print(f'Ошибка при загрузке файла: {e}')
        return False

def spec4EKF(elements):
    if not elements or elements == 1:
        return 1
    try:
        for element in elements:
            href = element.get_attribute('href')
            if href != "https://ekfgroup.com/ru/catalog":
                return href
    except Exception as e:
        print(f"Ошибка в spec4EKF: {e}")
    return None

# Основной код
excel_values = read_excel_column(excel_file_path, excel_column_name)
count_not = 0
count_yes = 0
cleanExcel(r"\\1cdbsrv\SystemFiles\pictures\result\Succes_EKF.xlsx")
cleanExcel(r"\\1cdbsrv\SystemFiles\pictures\result\Failed_EKF.xlsx")

for value in excel_values:
    url_tovar = GetPage(value)
    href = spec4EKF(url_tovar)
    print(href)
    if href == 1:
        count_not += 1
        listToExcel(value, "Fail")
    else:
        url_pic = GetPicture(href)
        if url_pic:
            name_file = value + ".png"
            if download(url_pic, name_file, href):
                count_yes += 1
                listToExcel(value, "Success")
            else:
                listToExcel(value, "Fail")
        else:
            listToExcel(value, "Fail")
            count_not += 1

print(f"Count_not = {count_not}")
print(f"Count_yes = {count_yes}")

driver.quit()
