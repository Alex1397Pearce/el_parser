import os
import random
import time
import requests
import pandas as pd
import openpyxl
from selenium import webdriver
from selenium.common import TimeoutException, NoSuchElementException, StaleElementReferenceException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

# Global variables
EXCEL_FILE_PATH = 'neox.xlsx'
EXCEL_COLUMN_NAME = 'Артикул'
URL_SEARCH = 'https://neox-group.ru/search/?q='

# Setup Chrome options
chrome_options = Options()
# chrome_options.add_argument('--headless')
driver = webdriver.Chrome(options=chrome_options)


def read_excel_column(file_path, column_name):
    df = pd.read_excel(file_path)
    return df[column_name].tolist()


def is_url(input_string):
    try:
        response = requests.get(input_string)
        return response.status_code == 200
    except requests.RequestException:
        return False


def get_page(value):
    url = URL_SEARCH + value
    try:
        driver.get(url)
        parent_element = WebDriverWait(driver, 5).until(
            EC.visibility_of_element_located((By.XPATH, '//div[@class="search-item"]/h4/a/b/ancestor::a'))
        )
        src_value = parent_element.get_attribute('href')
        time.sleep(random.uniform(2, 6))
        return src_value

    except (TimeoutException, NoSuchElementException, StaleElementReferenceException, WebDriverException):
        print(f'Error finding article {value}. URL not found.')
        return None


def get_picture(url):
    try:
        driver.get(url)
        element = WebDriverWait(driver, 10).until(
            EC.presence_of_element_located((By.CLASS_NAME, 'fancy'))
        )
        src_value = element.get_attribute('href')
        time.sleep(random.uniform(2, 6))
        return src_value
    except TimeoutException:
        print(f"Timeout error in GetPicture() for URL: {url}")
        return None


def clean_excel(file_path):
    if os.path.exists(file_path):
        os.remove(file_path)
    else:
        print(f"File {file_path} not found.")


def list_to_excel(value, status, path=""):
    file_path = r"\\1cdbsrv\SystemFiles\pictures\result\Succes_Neox.xlsx" if status == "Success" else r"\\1cdbsrv\SystemFiles\pictures\result\Failed_Neox.xlsx"

    if not os.path.exists(file_path):
        workbook = openpyxl.Workbook()
        workbook.save(file_path)

    workbook = openpyxl.load_workbook(file_path)
    while len(workbook.sheetnames) > 1:
        workbook.remove(workbook[workbook.sheetnames[1]])

    if not workbook.sheetnames:
        workbook.create_sheet()

    sheet = workbook.active
    first_empty_row = 1
    while sheet.cell(row=first_empty_row, column=1).value is not None:
        first_empty_row += 1

    sheet.cell(row=first_empty_row, column=1, value=value)
    if status == "Success":
        sheet.cell(row=first_empty_row, column=2, value=path)

    workbook.save(file_path)


def download(url, name_file, comment):
    try:
        response = requests.get(url)
        if response.status_code == 200:
            with open(fr"\\1cdbsrv\SystemFiles\pictures\{name_file}", 'wb') as file:
                file.write(response.content)
            print(f'File {name_file} successfully downloaded. URL: {comment}')
            return name_file
        else:
            print(f"Failed to download file. Status code: {response.status_code}")
            return False
    except requests.RequestException as e:
        print(f"Error downloading the file: {e}")
        return False


def main():
    # excel_values = ['4690612038179']
    excel_values = read_excel_column(EXCEL_FILE_PATH, EXCEL_COLUMN_NAME)
    count_not = 0
    count_yes = 0

    clean_excel(r"\\1cdbsrv\SystemFiles\pictures\result\Succes_Neox.xlsx")
    clean_excel(r"\\1cdbsrv\SystemFiles\pictures\result\Failed_Neox.xlsx")

    for value in excel_values:
        value = str(int(value))  # Ensure the value is a string
        url_tovar = get_page(value)
        if url_tovar:
            url_pic = get_picture(url_tovar)
            if url_pic:
                name_file = f"{value}.png"
                path = download(url_pic, name_file, url_tovar)
                if path:
                    path = "\\\\1cdbsrv\\SystemFiles\\pictures\\" + path
                    list_to_excel(value, "Success", path)
                    count_yes += 1
                else:
                    list_to_excel(value, "Fail")
                    count_not += 1
            else:
                list_to_excel(value, "Fail")
                count_not += 1
        else:
            list_to_excel(value, "Fail")
            count_not += 1

    print(f"Count_not = {count_not}")
    print(f"Count_yes = {count_yes}")

    driver.quit()


if __name__ == "__main__":
    main()

#     parent_element = WebDriverWait(driver, 5).until(
#         EC.visibility_of_element_located((By.XPATH, '//div.search-item/h4/a'))
#     )
#     src_value = parent_element.get_attribute('href')
#     time.sleep(random.uniform(2, 6))
#     return src_value
#
# except (TimeoutException, NoSuchElementException, StaleElementReferenceException, WebDriverException):
# print(f'Error finding article {value}. URL not found.')
# return None
#
#
# def get_picture(url):
#     try:
#         driver.get(url)
#         element = WebDriverWait(driver, 10).until(
#             EC.presence_of_element_located((By.CLASS_NAME, 'fancy'))