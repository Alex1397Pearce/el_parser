import pandas as pd

from selenium import webdriver
from selenium.common import TimeoutException, NoSuchElementException, StaleElementReferenceException, WebDriverException
from selenium.webdriver.chrome.options import Options
from selenium.webdriver.common.by import By
from selenium.webdriver.support.ui import WebDriverWait
from selenium.webdriver.support import expected_conditions as EC

import requests
import random
import time
import openpyxl
import os


# Объявление глобальных переменных
excel_file_path = 'Geniled.xlsx'
excel_column_name = 'Артикул'
url_search = 'https://geniled.ru/catalog/?q='

chrome_options = Options()
chrome_options.add_argument('--headless')
driver = webdriver.Chrome(options=chrome_options)

def read_excel_column(file_path, column_name):
    # Чтение определенного столбца из Excel файла и возвращение его в виде списка
    df = pd.read_excel(file_path)
    column_values = df[column_name].tolist()
    return column_values

def is_url(input_string):
    try:
        result = requests.get(input_string)
        return result.status_code
    except ValueError:
        return False

def CheckWoltaPage():
    try:
        check = driver.find_element(By.CLASS_NAME, "notetext")
        if check != None:
            return False
    except:
        return True


def GetPage(value):
    # code
    try:
        url = f"{url_search}%27{value}%27"
        driver.get(url)
        time.sleep(5)
        elements = driver.find_elements(By.CLASS_NAME, 'list_item_wrapp')
        for element in elements:
            art_element = element.find_element(By.CLASS_NAME, 'article_block')
            link_element = element.find_element(By.CLASS_NAME, 'dark_link')
            link = link_element.get_attribute('href')
            clean_art = art_element.text.split('Арт.: ')
            if clean_art[1] == value:
                return link


    except TimeoutException as e:
        print(f'Произошла ошибка ожидания')
        print(f'Артикул {value} не найден')
        return 1

    except NoSuchElementException as e:
        print(f'Элемент не был найден')
        print(f'Артикул {value} не найден')
        return 1

    except StaleElementReferenceException as e:
        print(f'Произошла ошибка с устаревшим элементом')
        print(f'Артикул {value} не найден')
        return 1

    except WebDriverException as e:
        print(f'Произошла общая ошибка WebDriver')
        print(f'Артикул {value} не найден')
        return 1


def GetPicture(url):
    try:
        # driver.get('https://tdme.ru/product/RM0109-0053.jpg')
        driver.get(url)
        parent_element = driver.find_element(By.CLASS_NAME, 'product-detail-gallery__slider--big')
        element = parent_element.find_element(By.TAG_NAME, 'img')
        src_value = element.get_attribute('src')
        print(f'Значение атрибута src: {src_value}')

        wait = random.randrange(2, 6)
        time.sleep(wait)
        return src_value
    except TimeoutException as e:
        print(f"Ошибка в процедуре GetPicture(). src_value = {url}")
        print(e)

def cleanExcel(file_path):
    try:
        os.remove(file_path)
    except FileNotFoundError:
        print(f"Файл {file_path} не найден.")

def listToExcel(value, status):
    if status == "Success":
        file_path = r"\\1cdbsrv\SystemFiles\pictures\result\Succes_Geniled.xlsx"
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            workbook.save(file_path)

            # Открываем существующий файл Excel
        workbook = openpyxl.load_workbook(file_path)

        # Если в файле больше одного листа, удаляем остальные
        while len(workbook.sheetnames) > 1:
            workbook.remove(workbook[workbook.sheetnames[1]])

        # Если в файле нет листов, создаем новый
        if not workbook.sheetnames:
            workbook.create_sheet()

        # Выбираем единственный лист
        sheet = workbook.active

        # Находим первую свободную ячейку в первом столбце
        first_empty_row = 1
        while sheet.cell(row=first_empty_row, column=1).value is not None:
            first_empty_row += 1

        # Добавляем данные в первую свободную ячейку
        sheet.cell(row=first_empty_row, column=1, value=value)

        # Сохраняем изменения в файл
        workbook.save(file_path)
    else:
        file_path = r"\\1cdbsrv\SystemFiles\pictures\result\Failed_Geniled.xlsx"
        if not os.path.exists(file_path):
            workbook = openpyxl.Workbook()
            workbook.save(file_path)

            # Открываем существующий файл Excel
        workbook = openpyxl.load_workbook(file_path)

        # Если в файле больше одного листа, удаляем остальные
        while len(workbook.sheetnames) > 1:
            workbook.remove(workbook[workbook.sheetnames[1]])

        # Если в файле нет листов, создаем новый
        if not workbook.sheetnames:
            workbook.create_sheet()

        # Выбираем единственный лист
        sheet = workbook.active

        # Находим первую свободную ячейку в первом столбце
        first_empty_row = 1
        while sheet.cell(row=first_empty_row, column=1).value is not None:
            first_empty_row += 1

        # Добавляем данные в первую свободную ячейку
        sheet.cell(row=first_empty_row, column=1, value=value)

        # Сохраняем изменения в файл
        workbook.save(file_path)



def download(url, name_file, comment):
    # Загрузка файла по URL
    response = requests.get(url)
    name_file = fr"\\1cdbsrv\SystemFiles\pictures\{name_file}"
    # Проверка успешности запроса
    if response.status_code == 200:
        # Сохранение файла на диск
        with open(name_file, 'wb') as file:
            file.write(response.content)
        print(f'Файл {name_file} успешно загружен. Ссылка на товар - {comment}')
        return True
    else:
        print(f'Ошибка при загрузке файла. Статус код: {response.status_code}')
        return False


# временные переменные


# Основной код
# excel_values = ['08168_12','08754','08755']
excel_values = read_excel_column(excel_file_path, excel_column_name)
count_not = 0
count_yes = 0
cleanExcel(file_path=r"\\1cdbsrv\SystemFiles\pictures\result\Succes_Geniled.xlsx")
cleanExcel(file_path=r"\\1cdbsrv\SystemFiles\pictures\result\Failed_Geniled.xlsx")

for value in excel_values:
    url_tovar = GetPage(value)
    print(url_tovar)
    if url_tovar == 1 or url_tovar == None:
        count_not = count_not + 1
        listToExcel(value, "Fail")
    else:
        url_pic = GetPicture(url_tovar)
        print(url_pic)
        name_file = value + ".png"
        download(url_pic, name_file, url_tovar)
        count_yes = count_yes + 1
        listToExcel(value, "Success")
print(f"Count_not = {count_not}")
print(f"Count_yes = {count_yes}")


driver.quit()