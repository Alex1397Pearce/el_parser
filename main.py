import os
import pandas as pd
import openpyxl
import openpyxl.styles.numbers
import requests
import time
from bs4 import BeautifulSoup


# link on web site IEK https://www.iek.ru/products/catalog/search?q=FP-V20-0-10-1-K10

class URLIterator:

    def __init__(self, items, base_url):
        self.items = items
        self.base_url = base_url
        self.index = 0

    def __iter__(self):
        return self

    def __next__(self):
        if self.index >= len(self.items):
            raise StopIteration

        item = self.items[self.index]
        url = f"{self.base_url}{item}"
        self.index += 1
        return url, item


class Files:

    def __init__(self, filepath):
        self.filepath = filepath
        self.status = self.file_exist()

    def file_exist(self):
        if os.path.exists(self.filepath):
            return True
        else:
            return False


class Reader(Files):
    def get_list(self, column_name="Артикул"):
        # check file exist
        try:
            df = pd.read_excel(self.filepath)
            column_values = df[column_name].tolist()
            return column_values
        except Exception as e:
            print(f"Error reading Excel file: {e}")
            return []


class Excel(Files):
    def __init__(self, filepath):
        super().__init__(filepath)
        if self.status:
            os.remove(self.filepath)
            workbook = openpyxl.Workbook()
            workbook.save(self.filepath)
        else:
            workbook = openpyxl.Workbook()
            workbook.save(self.filepath)

    def clean(self):
        if not self.status:
            pass
        os.remove(self.filepath)

    def list_to_excel(self, value_articul, url=""):
        workbook = openpyxl.load_workbook(self.filepath)
        while len(workbook.sheetnames) > 1:
            workbook.remove(workbook[workbook.sheetnames[1]])

        if not workbook.sheetnames:
            workbook.create_sheet()

        sheet = workbook.active
        first_empty_row = 1
        while sheet.cell(row=first_empty_row, column=1).value is not None:
            first_empty_row += 1
        sheet.cell(row=first_empty_row, column=1).number_format = openpyxl.styles.numbers.BUILTIN_FORMATS[1]
        sheet.cell(row=first_empty_row, column=1, value=value_articul)
        sheet.cell(row=first_empty_row, column=2, value=url)
        workbook.save(self.filepath)
        print(f"Записано в {self.filepath}: {value_articul} {url}")


class Browser:

    @staticmethod
    def get_page(url, timeout = 5, try_count = 5):
        while try_count > 0:
            response = requests.get(url)
            if response.status_code == 200:
                print(f"Загружена страница: {url}")
                return response.text
            else:
                time.sleep(timeout)
                try_count -= 1
                if try_count == 0:
                    response.raise_for_status()



    @staticmethod
    def download(url, name_file):
        response = requests.get(url)
        response.raise_for_status()
        print(f"Загружен файл с {url} в {name_file}")
        with open(name_file, 'wb') as file:
            file.write(response.content)



class Parser:

    def __init__(self, base_url):
        self.base_url = base_url

    @staticmethod
    def check_element(content_page, type_element, name_class):
        soup = BeautifulSoup(content_page, 'html.parser')
        if soup.find(type_element, class_=name_class):
            return True
        else:
            return False

    @staticmethod
    def get_attr_4el_by_class(page, type_element, class_name, attr_name):
        soup = BeautifulSoup(page, 'html.parser')
        element = soup.select_one(f"{type_element}.{class_name}")
        return element.attrs[attr_name]

    @staticmethod
    def get_attr_4el_by_id(page, id_name, attr_name):
        soup = BeautifulSoup(page, 'html.parser')
        element = soup.find(id=id_name).select_one("a.popup_link")
        return element.attrs[attr_name]

    @staticmethod
    def join_base_url(func):
        def wrapper(*args, **kwargs):
            self_instance = args[0]
            original_func = func(*args, **kwargs)
            modif_func = ''.join(f"{self_instance.base_url}{original_func}")
            return modif_func
        return wrapper

    # @join_base_urlurl
    def test_func(self):
        return "/text"

    def check_element_ref(self, type_element, name_class):
        soup = BeautifulSoup(self.search_page, 'html.parser')
        if soup.find(type_element, class_=name_class):
            return True
        else:
            return False

    def get_element_by_id(self, id):
        soup = BeautifulSoup(self.search_page, 'html.parser')
        element = soup.find(id=id).select_one("a.popup_link")
        return element.attrs['href']

    def get_element(self, type_element, name_class, name_attr):
        soup = BeautifulSoup(self.search_page, 'html.parser')
        element = soup.select_one(f"{type_element}.{name_class}")
        return element.attrs[name_attr]

    def get_link_element(self, type_element, name_class, name_attr='href'):
        element = self.get_element(type_element, name_class, name_attr)
        return f"{self.base_url}{element}"

    def get_element_from(self, page, type_element, name_class, name_attr='href'):
        soup = BeautifulSoup(page, 'html.parser')
        element = soup.select_one(f"{type_element}.{name_class}")
        return element.attrs[name_attr]
